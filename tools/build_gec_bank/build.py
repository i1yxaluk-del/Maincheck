"""
Extract GEC pairs from open Russian corpora and build extended GEC bank.

Sources:
  - LORuGEC: https://github.com/ReginaNasyrova/LORuGEC
    960 (wrong, right) pairs across 48 rules of Russian grammar.
    Released as part of Sorokin & Nasyrova, BEA @ ACL 2025.

Usage:
  # 1. Скачать исходник (см. tools/build_gec_bank/README.md):
  #    git clone --depth 1 https://github.com/ReginaNasyrova/LORuGEC.git \
  #      tools/build_gec_bank/sources/LORuGEC
  # 2. Сборка:
  python3 tools/build_gec_bank/build.py

Output:
  server/shared/gec_seed/gec_bank_extended.jsonl
    — пары LORuGEC, не пересекающиеся с уже существующим gec_bank.jsonl.
"""
import argparse
import hashlib
import json
import re
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
# Поддерживаем оба варианта расположения: либо xlsx прямо в sources/, либо
# полный клон репозитория ReginaNasyrova/LORuGEC в sources/LORuGEC/.
_SOURCES = REPO_ROOT / "tools/build_gec_bank/sources"
_LORUGEC_CANDIDATES = [
    _SOURCES / "LORuGEC.xlsx",
    _SOURCES / "LORuGEC" / "LORuGEC.xlsx",
]
DEFAULT_EXISTING_JSONL = REPO_ROOT / "server/shared/gec_seed/gec_bank.jsonl"
DEFAULT_OUTPUT = REPO_ROOT / "server/shared/gec_seed/gec_bank_extended.jsonl"


def _find_lorugec(explicit: Path | None) -> Path | None:
    if explicit is not None:
        return explicit if explicit.exists() else None
    for cand in _LORUGEC_CANDIDATES:
        if cand.exists():
            return cand
    return None


def detokenize_ru(text: str) -> str:
    """Collapse whitespace around Russian/English punctuation."""
    # Remove space before closing puncts: ,.!?;:)
    text = re.sub(r"\s+([,.!?;:)\]…»])", r"\1", text)
    # Remove space after opening puncts: ([«
    text = re.sub(r"([(\[«])\s+", r"\1", text)
    # Collapse multiple spaces
    text = re.sub(r"\s+", " ", text).strip()
    # Em-dash with surrounding spaces stays
    return text


def has_nested_quotes(text: str) -> bool:
    """Check for nested «...«...»...» which the model often hallucinates."""
    open_count = 0
    for ch in text:
        if ch == "«":
            open_count += 1
            if open_count > 1:
                return True
        elif ch == "»":
            open_count = max(0, open_count - 1)
    return False


def normalize_pair_key(wrong: str, right: str) -> str:
    """Hash of normalized pair for dedup."""
    norm = (wrong.strip().lower(), right.strip().lower())
    return hashlib.sha256(repr(norm).encode("utf-8")).hexdigest()


def load_existing(path: Path) -> set[str]:
    keys = set()
    if not path.exists():
        return keys
    with path.open() as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            d = json.loads(line)
            keys.add(normalize_pair_key(d["wrong"], d["right"]))
    return keys


def extract_lorugec(xlsx_path: Path) -> list[dict]:
    """Read LORuGEC.xlsx → list of pair dicts."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}

    pairs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Initial sentence"]] is None:
            continue
        wrong = detokenize_ru(str(row[idx["Initial sentence"]]))
        right = detokenize_ru(str(row[idx["Correct sentence"]]))
        same = str(row[idx["Are both sentences the same?"]] or "").strip().lower()
        if same == "yes":
            continue
        if not wrong or not right or wrong == right:
            continue
        rule = str(row[idx["The rule"]] or "").strip()
        definition = str(row[idx["The definition of the rule"]] or "").strip()
        section = str(row[idx["Grammar section"]] or "").strip()
        pairs.append({
            "wrong": wrong,
            "right": right,
            "rule": rule,
            "definition": definition,
            "section": section,
        })
    return pairs


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--lorugec", type=Path, default=None,
                        help="Путь к LORuGEC.xlsx (по умолчанию ищется в sources/).")
    parser.add_argument("--existing", type=Path, default=DEFAULT_EXISTING_JSONL)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    lorugec_xlsx = _find_lorugec(args.lorugec)
    if lorugec_xlsx is None:
        raise SystemExit(
            "LORuGEC.xlsx не найден.\n"
            "Скачайте исходник через:\n"
            "  git clone --depth 1 https://github.com/ReginaNasyrova/LORuGEC.git \\\n"
            f"    {_SOURCES}/LORuGEC\n"
            "или укажите путь явно через --lorugec /path/to/LORuGEC.xlsx"
        )

    existing_keys = load_existing(args.existing)
    print(f"Existing bank: {len(existing_keys)} pairs")

    print(f"Reading LORuGEC from {lorugec_xlsx}")
    pairs = extract_lorugec(lorugec_xlsx)
    print(f"LORuGEC raw pairs: {len(pairs)}")

    # Filter: nested quotes, dedup against existing, dedup within
    seen = set(existing_keys)
    out = []
    skipped_nested = 0
    skipped_dup = 0
    for p in pairs:
        if has_nested_quotes(p["wrong"]) or has_nested_quotes(p["right"]):
            skipped_nested += 1
            continue
        key = normalize_pair_key(p["wrong"], p["right"])
        if key in seen:
            skipped_dup += 1
            continue
        seen.add(key)
        out.append(p)

    print(f"Skipped nested quotes: {skipped_nested}")
    print(f"Skipped duplicates of existing bank: {skipped_dup}")
    print(f"Output pairs: {len(out)}")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8") as f:
        for d in out:
            f.write(json.dumps(d, ensure_ascii=False) + "\n")
    print(f"Written to {args.output}")

    # Stats by section
    from collections import Counter
    sec_counts = Counter(p["section"] for p in out)
    rule_counts = Counter(p["rule"] for p in out)
    print("\nBy section:")
    for s, c in sec_counts.most_common():
        print(f"  {c:4d}  {s}")
    print(f"\nUnique rules: {len(rule_counts)}")
    print(f"Top rules:")
    for r, c in rule_counts.most_common(5):
        print(f"  {c:4d}  {r[:80]}")


if __name__ == "__main__":
    main()
